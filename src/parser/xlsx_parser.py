"""
Owned by: Pooja
Initial code provided by: Fatima

xlsx_parser.py — parses XLSX log files using openpyxl.
Per Section 4.7.1 step 2: 'IF format == "xlsx": Parse using openpyxl'.
Consumed by: log_parser.py
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from src.parser.csv_parser import MalformedFileError

def parse_xlsx(file_path: str) -> list[dict]:
    """Reads an XLSX log file and returns a list of raw row dicts.

    Raises:
        FileNotFoundError: if file_path does not exist.
        MalformedFileError: if the file is not a valid XLSX workbook or has no header.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"XLSX file not found: {file_path}")

    try:
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    except InvalidFileException as exc:
        raise MalformedFileError(
            f"XLSX file is not a valid workbook: {file_path} — {exc}"
        ) from exc
    except Exception as exc:
        raise MalformedFileError(
            f"XLSX file could not be opened: {file_path} — {exc}"
        ) from exc

    worksheet = workbook.active
    rows: list[dict] = []
    row_iter = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(row_iter)
    except StopIteration:
        workbook.close()
        raise MalformedFileError(f"XLSX file is empty (no rows at all): {file_path}")

    headers = [
        str(h) if h is not None else f"column_{i}"
        for i, h in enumerate(header_row)
    ]

    if all(h.startswith("column_") for h in headers):
        workbook.close()
        raise MalformedFileError(
            f"XLSX file has no recognisable header row (all cells empty): {file_path}"
        )

    for raw_row in row_iter:
        if all(cell is None for cell in raw_row):
            continue
        row_dict = {
            header: (
                "" if (col_index >= len(raw_row) or raw_row[col_index] is None)
                else str(raw_row[col_index])
            )
            for col_index, header in enumerate(headers)
        }
        rows.append(row_dict)

    workbook.close()
    return rows